from django.shortcuts import render, redirect
from django.contrib import messages
from django.db.models import Sum
from datetime import datetime
from ..models import Famille, Revenu, Depense, Budget
from ..forms import RechercheForm
import csv
import openpyxl
from django.http import HttpResponse

def accueil(request):
    """Page d'accueil simple"""
    return render(request, 'gestionbudgetfamille10/accueil.html')

def dashboard(request):
    """Dashboard principal avec les indicateurs clés"""
    # Récupérer la famille sélectionnée (pour l'instant la première)
    famille = Famille.objects.first()
    
    if not famille:
        messages.warning(request, "Vous devez d'abord créer une famille.")
        return redirect('ajouter_famille')
    
    # Date courante
    aujourdhui = datetime.now()
    mois_courant = aujourdhui.month
    annee_courante = aujourdhui.year
    
    # Calculs pour le mois en cours
    revenus_mois = Revenu.objects.filter(
        Date__month=mois_courant,
        Date__year=annee_courante,
        idMembre__idFamille=famille
    ).aggregate(total=Sum('Montant'))['total'] or 0
    
    depenses_mois = Depense.objects.filter(
        Date__month=mois_courant,
        Date__year=annee_courante,
        idMembre__idFamille=famille
    ).aggregate(total=Sum('Montant'))['total'] or 0
    
    # Budget du mois
    budget_mois = Budget.objects.filter(
        Mois=mois_courant,
        Annee=annee_courante,
        idFamille=famille
    ).first()
    
    # Alertes de dépassement
    alertes = []
    if budget_mois and depenses_mois > budget_mois.MontantPrevu:
        alertes.append({
            'type': 'danger',
            'message': f"Dépassement du budget de {depenses_mois - budget_mois.MontantPrevu}CFA !"
        })
    elif budget_mois and depenses_mois > budget_mois.MontantPrevu * 0.9:
        alertes.append({
            'type': 'warning',
            'message': f"Attention : vous avez utilisé {depenses_mois/budget_mois.MontantPrevu*100:.1f}% de votre budget."
        })
    
    # Dernières opérations
    dernieres_depenses = Depense.objects.filter(
        idMembre__idFamille=famille
    ).order_by('-Date')[:5]
    
    derniers_revenus = Revenu.objects.filter(
        idMembre__idFamille=famille
    ).order_by('-Date')[:5]
    
    context = {
        'famille': famille,
        'revenus_mois': revenus_mois,
        'depenses_mois': depenses_mois,
        'solde_mois': revenus_mois - depenses_mois,
        'budget_mois': budget_mois,
        'alertes': alertes,
        'dernieres_depenses': dernieres_depenses,
        'derniers_revenus': derniers_revenus,
        'mois_courant': mois_courant,
        'annee_courante': annee_courante,
    }
    return render(request, 'gestionbudgetfamille10/dashboard.html', context)

def alertes(request):
    """Page des alertes"""
    famille = Famille.objects.first()
    if not famille:
        return redirect('ajouter_famille')
    
    # Vérifier tous les budgets
    budgets = Budget.objects.filter(idFamille=famille)
    alertes_list = []
    
    for budget in budgets:
        depenses = Depense.objects.filter(
            Date__month=budget.Mois,
            Date__year=budget.Annee,
            idMembre__idFamille=famille
        ).aggregate(total=Sum('Montant'))['total'] or 0
        
        if depenses > budget.MontantPrevu:
            alertes_list.append({
                'budget': budget,
                'depenses': depenses,
                'depassement': depenses - budget.MontantPrevu,
                'pourcentage': (depenses / budget.MontantPrevu) * 100
            })
    
    context = {
        'famille': famille,
        'alertes': alertes_list
    }
    return render(request, 'gestionbudgetfamille10/alertes.html', context)

def import_excel(request):
    """Vue pour l'import Excel/CSV"""
    if request.method == 'POST':
        fichier = request.FILES.get('fichier')
        type_import = request.POST.get('type')
        
        if not fichier:
            messages.error(request, "Veuillez sélectionner un fichier.")
            return redirect('import_excel')
        
        # Logique d'import selon le type
        if fichier.name.endswith('.csv'):
            # Traitement CSV
            decoded_file = fichier.read().decode('utf-8').splitlines()
            reader = csv.reader(decoded_file)
            next(reader)  # Sauter l'en-tête
            
            for row in reader:
                if type_import == 'revenus':
                    # Traitement des revenus
                    pass
                elif type_import == 'depenses':
                    # Traitement des dépenses
                    pass
                    
        elif fichier.name.endswith(('.xlsx', '.xls')):
            # Traitement Excel
            wb = openpyxl.load_workbook(fichier)
            sheet = wb.active
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if type_import == 'revenus':
                    # Traitement des revenus
                    pass
                # etc.
        
        messages.success(request, f"Import {type_import} réussi !")
        return redirect('dashboard')
    
    return render(request, 'gestionbudgetfamille10/import.html')

def export_excel(request, modele):
    """Export des données en Excel"""
    famille = Famille.objects.first()
    
    if modele == 'revenus':
        queryset = Revenu.objects.filter(idMembre__idFamille=famille)
        filename = "revenus"
    elif modele == 'depenses':
        queryset = Depense.objects.filter(idMembre__idFamille=famille)
        filename = "depenses"
    else:
        messages.error(request, "Modèle non valide")
        return redirect('dashboard')
    
    # Création du fichier Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename.capitalize()
    
    # En-têtes
    if modele == 'revenus':
        ws.append(['Date', 'Membre', 'Type', 'Montant'])
        for obj in queryset:
            ws.append([
                obj.Date.strftime('%d/%m/%Y'),
                str(obj.idMembre),
                obj.idTypeRevenu.NomType,
                float(obj.Montant)
            ])
    else:
        ws.append(['Date', 'Membre', 'Catégorie', 'Mode paiement', 'Montant'])
        for obj in queryset:
            ws.append([
                obj.Date.strftime('%d/%m/%Y'),
                str(obj.idMembre),
                obj.idCategorieDepense.NomCategorie,
                obj.idModePaiement.NomMode,
                float(obj.Montant)
            ])
    
    wb.save(response)
    return response